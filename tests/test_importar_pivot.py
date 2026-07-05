"""Tests del comando importar_pivot contra un PIVOT sintético (P16, O13: sin archivos reales)."""

from pathlib import Path

import openpyxl
import pytest
from django.core.management import call_command
from django.core.management.base import CommandError

from apps.ops.models import EjecucionPipeline
from apps.perfiles.models import PalabraIntencion, PerfilFiltro, ReglaKeyword

pytestmark = [pytest.mark.integration, pytest.mark.django_db]

# La fila 6 es donde empiezan los datos en las hojas de filtros del PIVOT original.
FILA_DATOS = 6


def crear_pivot_sintetico(ruta: Path) -> Path:
    """PIVOT mínimo con el layout real: catálogo, intención global y una hoja de filtros."""
    wb = openpyxl.Workbook()

    ws_equipos = wb.active
    assert ws_equipos is not None
    ws_equipos.title = "00-Equipos"
    ws_equipos.append(["codigo", "nombre", "hoja_filtros", "descripcion", "activo"])
    ws_equipos.append(["TELECOM", "Telecomunicaciones", "06-Telecom", "Equipo demo", "TRUE"])

    ws_intencion = wb.create_sheet("01-Intencion_Global")
    ws_intencion.append(["intencion_requerida", "intencion_vetada"])
    ws_intencion.append(["ingeniería", "adquisición de"])
    ws_intencion.append(["consultoría", None])

    ws_filtros = wb.create_sheet("06-Telecom")
    # Columnas: 0 incluir/nombre, 4 excluir/nombre, 12 bypass, 13 exclusión dura.
    ws_filtros.cell(row=FILA_DATOS, column=1, value="fibra óptica")
    ws_filtros.cell(row=FILA_DATOS, column=5, value="pavimentación")
    ws_filtros.cell(row=FILA_DATOS, column=13, value="ITS")
    ws_filtros.cell(row=FILA_DATOS, column=14, value="alimentación")
    # Duplicado deliberado: el import debe deduplicar.
    ws_filtros.cell(row=FILA_DATOS + 1, column=1, value="FIBRA OPTICA")

    wb.save(ruta)
    return ruta


def test_importa_perfiles_reglas_e_intencion(tmp_path: Path) -> None:
    pivot = crear_pivot_sintetico(tmp_path / "pivot.xlsx")

    call_command("importar_pivot", str(pivot))

    perfil = PerfilFiltro.objects.get(codigo="TELECOM")
    assert perfil.activo is True
    textos = set(perfil.reglas.values_list("texto", flat=True))
    # Normalizado y deduplicado: 'fibra óptica' y 'FIBRA OPTICA' son una sola regla.
    assert textos == {"FIBRA OPTICA", "PAVIMENTACION", "ITS", "ALIMENTACION"}
    assert perfil.reglas.filter(tipo=ReglaKeyword.Tipo.BYPASS, texto="ITS").exists()
    assert PalabraIntencion.objects.filter(
        tipo=PalabraIntencion.Tipo.VETADA, texto="ADQUISICION DE"
    ).exists()


def test_reimportar_es_idempotente(tmp_path: Path) -> None:
    pivot = crear_pivot_sintetico(tmp_path / "pivot.xlsx")
    call_command("importar_pivot", str(pivot))
    total_reglas = ReglaKeyword.objects.count()

    call_command("importar_pivot", str(pivot))

    assert ReglaKeyword.objects.count() == total_reglas
    assert PerfilFiltro.objects.count() == 1


def test_dry_run_no_escribe(tmp_path: Path) -> None:
    pivot = crear_pivot_sintetico(tmp_path / "pivot.xlsx")

    call_command("importar_pivot", str(pivot), "--dry-run")

    assert PerfilFiltro.objects.count() == 0
    assert ReglaKeyword.objects.count() == 0
    # La observabilidad se conserva incluso en dry-run (O5).
    ejecucion = EjecucionPipeline.objects.get()
    assert ejecucion.estado == EjecucionPipeline.Estado.EXITOSA
    assert ejecucion.metricas["dry_run"] is True


def test_pivot_sin_hoja_de_equipos_falla_con_mensaje_accionable(tmp_path: Path) -> None:
    ruta = tmp_path / "roto.xlsx"
    wb = openpyxl.Workbook()
    wb.save(ruta)

    with pytest.raises(CommandError, match="00-Equipos"):
        call_command("importar_pivot", str(ruta))

    assert EjecucionPipeline.objects.get().estado == EjecucionPipeline.Estado.FALLIDA
