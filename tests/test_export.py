"""Tests del export a Excel: el builder y la acción del Admin."""

from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from django.test import Client

from apps.licitaciones.export import ENCABEZADOS, construir_excel
from apps.licitaciones.models import Licitacion
from tests.factories import LicitacionFactory, OrganismoFactory

pytestmark = [pytest.mark.integration, pytest.mark.django_db]


def test_construir_excel_contiene_encabezados_y_filas() -> None:
    organismo = OrganismoFactory(nombre="Ministerio Sintetico", region="Valparaíso")
    LicitacionFactory(codigo_externo="111-1-L126", organismo=organismo)
    LicitacionFactory(codigo_externo="222-2-L126")

    contenido = construir_excel(Licitacion.objects.select_related("organismo"))

    wb = openpyxl.load_workbook(BytesIO(contenido))
    ws = wb.active
    assert ws is not None
    filas = list(ws.iter_rows(values_only=True))
    assert filas[0] == ENCABEZADOS
    codigos = {fila[0] for fila in filas[1:]}
    assert codigos == {"111-1-L126", "222-2-L126"}
    organismos = {fila[2] for fila in filas[1:]}
    assert "Ministerio Sintetico" in organismos


def test_accion_admin_descarga_xlsx() -> None:
    admin_user = get_user_model().objects.create_superuser(
        username="admin", password="clave!segura1"
    )
    licitacion = LicitacionFactory()
    cliente = Client()
    cliente.force_login(admin_user)

    respuesta = cliente.post(
        "/admin/licitaciones/licitacion/",
        {"action": "exportar_a_excel", "_selected_action": [licitacion.pk]},
    )

    assert respuesta.status_code == 200
    assert respuesta["Content-Type"].startswith("application/vnd.openxmlformats")
    assert ".xlsx" in respuesta["Content-Disposition"]
    wb = openpyxl.load_workbook(BytesIO(respuesta.content))
    ws = wb.active
    assert ws is not None
    assert ws.max_row == 2
