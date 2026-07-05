"""Adaptador BulkPortalSource: el Excel masivo diario del portal → filas de dominio.

Implementación del puerto FuenteLicitaciones (PLAN.md 2.1). Es el ÚNICO módulo del
sistema que conoce el formato XLSX del portal (P3): detecta el encabezado
dinámicamente (el portal a veces antepone filas de título), mapea columnas por
nombre normalizado y entrega dicts neutros. Si ChileCompra cambia el canal a JSON,
se reemplaza este módulo y nada más.
"""

import logging
import shutil
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl
import requests

from domain.normalizacion import limpiar_codigo_licitacion, normalizar_texto

logger = logging.getLogger(__name__)

# El portal antepone a veces filas de título; el encabezado real se busca en las primeras N.
MAX_FILAS_BUSQUEDA_HEADER = 10
# Columna cuya presencia identifica la fila de encabezado (mismo criterio que el original).
COLUMNA_REFERENCIA_HEADER = "NIVEL 1"
DESCARGA_TIMEOUT_SEGUNDOS = 120
DESCARGA_CHUNK_BYTES = 8192

# Nombre normalizado de columna del portal → clave neutra del dict de salida.
MAPEO_COLUMNAS_BULK: dict[str, str] = {
    "NUMERO ADQUISICION": "codigo_externo",
    "NOMBRE ADQUISICION": "nombre",
    "DESCRIPCION": "descripcion",
    "NIVEL 1": "nivel1",
    "NIVEL 2": "nivel2",
    "NIVEL 3": "nivel3",
    "GENERICO": "generico",
    "ORGANISMO": "organismo",
    "TIPO ADQUISICION": "tipo_adquisicion",
    "DESCRIPCION DEL PRODUCTO/SERVICIO": "descripcion_producto",
    "REGION": "region",
    "ESTADO": "estado_fuente",
    "MONEDA": "moneda",
    "MONTO ESTIMADO": "monto_estimado",
    "FECHA PUBLICACION": "fecha_publicacion",
    "FECHA CIERRE": "fecha_cierre",
}

# Sin estas columnas el motor de matching no puede operar (mismas exigencias que etapa 2).
COLUMNAS_MINIMAS = ("codigo_externo", "nombre", "nivel1", "organismo")


class FormatoBulkError(Exception):
    """El archivo del portal no tiene la estructura esperada. Mensaje accionable (O17)."""


@dataclass(frozen=True)
class FilaBulk:
    """Una licitación tal como viene en el bulk, ya con claves neutras.

    `datos` contiene los campos mapeados; `cruda` la fila completa original
    (termina en Licitacion.raw_bulk para auditoría y re-proceso).
    """

    codigo_externo: str
    datos: dict[str, Any]
    cruda: dict[str, Any]


def descargar_bulk(url: str, destino: Path) -> Path:
    """Descarga el Excel masivo del portal a `destino` (streaming, timeout explícito)."""
    logger.info("Descargando bulk del portal: %s", url)
    respuesta = requests.get(url, stream=True, timeout=DESCARGA_TIMEOUT_SEGUNDOS)
    respuesta.raise_for_status()
    destino.parent.mkdir(parents=True, exist_ok=True)
    with destino.open("wb") as archivo:
        for chunk in respuesta.iter_content(chunk_size=DESCARGA_CHUNK_BYTES):
            if chunk:
                archivo.write(chunk)
    tamano_mb = destino.stat().st_size / (1024 * 1024)
    logger.info("Descarga completa: %.2f MB en %s", tamano_mb, destino)
    return destino


# Un XLSX real es un ZIP que contiene este miembro; su ausencia delata un ZIP envolvente.
MIEMBRO_XLSX_VALIDO = "[Content_Types].xml"


def desempaquetar_bulk(ruta: Path) -> Path:
    """El portal a veces entrega el XLSX envuelto en un ZIP contenedor; lo extrae.

    Port de la lógica de la etapa 0 del proyecto original, incluido el guard de
    Zip Slip: del miembro se usa SOLO el nombre base, nunca rutas con directorios.
    Si el archivo ya es un XLSX válido (contiene '[Content_Types].xml'), se
    devuelve tal cual.
    """
    if not zipfile.is_zipfile(ruta):
        return ruta
    with zipfile.ZipFile(ruta) as zf:
        nombres = zf.namelist()
        if MIEMBRO_XLSX_VALIDO in nombres:
            return ruta
        candidatos = [n for n in nombres if n.lower().endswith((".xlsx", ".xls"))]
        if not candidatos:
            raise FormatoBulkError(
                f"La descarga del portal es un ZIP sin ningún Excel adentro "
                f"(miembros: {nombres[:5]}). El formato del portal cambió; revisa "
                f"el archivo en '{ruta}'."
            )
        miembro = candidatos[0]
        nombre_seguro = Path(miembro).name
        if not nombre_seguro:
            raise FormatoBulkError(f"Nombre de miembro ZIP inseguro o vacío: '{miembro}'.")
        destino = ruta.parent / nombre_seguro
        logger.info("Bulk envuelto en ZIP: extrayendo '%s'", nombre_seguro)
        with zf.open(miembro) as origen, destino.open("wb") as salida:
            shutil.copyfileobj(origen, salida)
        return destino


# ANN401 justificado: openpyxl no expone un tipo publico comun para Worksheet en
# modo normal y read-only; el contrato real es "algo con iter_rows".
def _detectar_header(ws: Any) -> tuple[int, dict[int, str]]:  # noqa: ANN401
    """Encuentra la fila de encabezado y devuelve (número de fila, índice→clave neutra)."""
    for numero_fila, fila in enumerate(
        ws.iter_rows(min_row=1, max_row=MAX_FILAS_BUSQUEDA_HEADER, values_only=True), start=1
    ):
        normalizadas = [normalizar_texto(celda) for celda in fila]
        if COLUMNA_REFERENCIA_HEADER in normalizadas:
            indice = {
                i: MAPEO_COLUMNAS_BULK[nombre]
                for i, nombre in enumerate(normalizadas)
                if nombre in MAPEO_COLUMNAS_BULK
            }
            return numero_fila, indice
    raise FormatoBulkError(
        f"No se encontró la columna '{COLUMNA_REFERENCIA_HEADER}' en las primeras "
        f"{MAX_FILAS_BUSQUEDA_HEADER} filas del archivo: el formato del portal cambió "
        f"o el archivo no es el bulk de licitaciones. Revisa el archivo manualmente."
    )


def leer_bulk(origen: Path | BinaryIO) -> list[FilaBulk]:
    """Parsea el Excel masivo y devuelve las filas con código válido.

    Filas sin código de licitación se descartan (el portal incluye subtotales y
    filas decorativas). La deduplicación por código la hace la BD vía upsert.
    """
    wb = openpyxl.load_workbook(origen, read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise FormatoBulkError("El archivo no tiene hojas legibles.")
        fila_header, indice = _detectar_header(ws)

        faltantes = set(COLUMNAS_MINIMAS) - set(indice.values())
        if faltantes:
            raise FormatoBulkError(
                f"El bulk no trae columnas mínimas para el filtrado: {sorted(faltantes)}. "
                f"Columnas reconocidas: {sorted(indice.values())}."
            )

        filas: list[FilaBulk] = []
        for fila in ws.iter_rows(min_row=fila_header + 1, values_only=True):
            datos = {clave: fila[i] if i < len(fila) else None for i, clave in indice.items()}
            codigo = limpiar_codigo_licitacion(datos.get("codigo_externo"))
            if not codigo:
                continue
            cruda = {
                f"col_{i}": _serializable(valor)
                for i, valor in enumerate(fila)
                if valor is not None
            }
            datos["codigo_externo"] = codigo
            filas.append(
                FilaBulk(
                    codigo_externo=codigo,
                    datos={k: _serializable(v) for k, v in datos.items()},
                    cruda=cruda,
                )
            )
        return filas
    finally:
        wb.close()


def _serializable(valor: Any) -> Any:  # noqa: ANN401 — celdas Excel: tipos heterogeneos por diseno
    """Convierte valores de celda a tipos JSON-compatibles (fechas → ISO 8601)."""
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    return valor
