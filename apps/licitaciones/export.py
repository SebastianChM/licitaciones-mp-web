"""Adaptador de salida: licitaciones -> Excel para el equipo comercial.

Reutiliza la competencia del proyecto original (reportes openpyxl) pero al
revés: allá el Excel era el almacenamiento; aquí es solo un formato de
exportación puntual que se genera desde la BD y no guarda estado.
"""

from datetime import datetime
from io import BytesIO

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.licitaciones.models import Licitacion

ENCABEZADOS = (
    "Código",
    "Nombre",
    "Organismo",
    "Región",
    "Estado",
    "Moneda",
    "Monto estimado",
    "Fecha publicación",
    "Fecha cierre",
    "URL ficha",
)
# Ancho de columna proporcional al contenido típico de cada campo.
ANCHOS = (16, 60, 40, 22, 14, 8, 16, 16, 18, 40)
_COLOR_ENCABEZADO = "1F4E79"


def construir_excel(queryset: QuerySet[Licitacion]) -> bytes:
    """Genera el XLSX en memoria. El queryset debe venir con select_related('organismo')."""
    wb = Workbook()
    # worksheets[0] y no .active: un Workbook nuevo siempre trae una hoja, y este
    # accesor está tipado como no-opcional (evita un assert en código productivo).
    ws = wb.worksheets[0]
    ws.title = "Licitaciones"

    ws.append(ENCABEZADOS)
    relleno = PatternFill(start_color=_COLOR_ENCABEZADO, fill_type="solid")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno
    for indice, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = ancho
    ws.freeze_panes = "A2"

    for licitacion in queryset:
        ws.append(
            (
                licitacion.codigo_externo,
                licitacion.nombre,
                licitacion.organismo.nombre if licitacion.organismo else "",
                licitacion.organismo.region if licitacion.organismo else "",
                licitacion.estado_fuente,
                licitacion.moneda,
                licitacion.monto_estimado,
                licitacion.fecha_publicacion,
                # Excel no acepta datetimes con timezone: se exporta naive local.
                licitacion.fecha_cierre.astimezone().replace(tzinfo=None)
                if licitacion.fecha_cierre
                else None,
                licitacion.url_ficha,
            )
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def nombre_archivo() -> str:
    return f"licitaciones_{datetime.now().astimezone():%Y%m%d_%H%M}.xlsx"
