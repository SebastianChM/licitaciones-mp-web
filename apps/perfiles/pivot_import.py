"""Adaptador de importación: PIVOT_MAESTRO.xlsx → modelos de perfiles.

Puente de migración (PLAN.md R5): lee el Excel maestro del proyecto original y
puebla PerfilFiltro/ReglaKeyword/PalabraIntencion. Es la ÚNICA pieza de esta app
que conoce el layout del PIVOT; el resto del sistema solo ve modelos.

Layout portado de core/profile_loader.py del proyecto original:
- Hoja '00-Equipos' (header fila 1): codigo | nombre | hoja_filtros | descripcion | activo
- Hoja '01-Intencion_Global' (opcional): intencion_requerida | intencion_vetada
- Hoja de filtros por equipo: datos desde la fila 6; columnas por posición:
  0-3 incluir (nombre, nivel1, nivel2, nivel3), 4-11 excluir (nombre, nivel1,
  nivel2, nivel3, generico, componente, organismo, valor), 12 bypass, 13 exclusión dura.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from apps.perfiles.models import PalabraIntencion, PerfilFiltro, ReglaKeyword
from domain.normalizacion import normalizar_texto

logger = logging.getLogger(__name__)

HOJA_EQUIPOS = "00-Equipos"
HOJA_INTENCION_GLOBAL = "01-Intencion_Global"
# Fila (1-indexed) donde empiezan los datos en las hojas de filtros del PIVOT.
FILA_INICIO_DATOS_FILTROS = 6

Campo = ReglaKeyword.Campo
Tipo = ReglaKeyword.Tipo

# Columna 0-indexed de la hoja de filtros → (tipo de regla, campo objetivo).
# El orden de columnas es contrato con el PIVOT original; no reordenar.
MAPEO_COLUMNAS: dict[int, tuple[str, str]] = {
    0: (Tipo.INCLUIR, Campo.NOMBRE),
    1: (Tipo.INCLUIR, Campo.NIVEL1),
    2: (Tipo.INCLUIR, Campo.NIVEL2),
    3: (Tipo.INCLUIR, Campo.NIVEL3),
    4: (Tipo.EXCLUIR, Campo.NOMBRE),
    5: (Tipo.EXCLUIR, Campo.NIVEL1),
    6: (Tipo.EXCLUIR, Campo.NIVEL2),
    7: (Tipo.EXCLUIR, Campo.NIVEL3),
    8: (Tipo.EXCLUIR, Campo.GENERICO),
    9: (Tipo.EXCLUIR, Campo.COMPONENTE),
    10: (Tipo.EXCLUIR, Campo.ORGANISMO),
    11: (Tipo.EXCLUIR, Campo.VALOR),
    12: (Tipo.BYPASS, ""),
    13: (Tipo.EXCLUSION_DURA, ""),
}


class PivotImportError(Exception):
    """Error de estructura del PIVOT, con mensaje accionable para el operador (O17)."""


@dataclass
class ResultadoImportacion:
    """Métricas de una importación, para el log y EjecucionPipeline."""

    perfiles_creados: int = 0
    perfiles_actualizados: int = 0
    reglas_creadas: int = 0
    reglas_existentes: int = 0
    intencion_creadas: int = 0
    advertencias: list[str] = field(default_factory=list)

    def como_dict(self) -> dict[str, object]:
        return {
            "perfiles_creados": self.perfiles_creados,
            "perfiles_actualizados": self.perfiles_actualizados,
            "reglas_creadas": self.reglas_creadas,
            "reglas_existentes": self.reglas_existentes,
            "intencion_creadas": self.intencion_creadas,
            "advertencias": self.advertencias,
        }


def _leer_equipos(wb: openpyxl.Workbook) -> list[dict[str, str | bool]]:
    """Lee el catálogo de equipos de la hoja 00-Equipos."""
    if HOJA_EQUIPOS not in wb.sheetnames:
        raise PivotImportError(
            f"Falta la hoja '{HOJA_EQUIPOS}' en el PIVOT: esa hoja define qué equipos existen. "
            f"Hojas encontradas: {wb.sheetnames}"
        )
    ws = wb[HOJA_EQUIPOS]
    filas = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower() for c in next(filas)]
    except StopIteration as e:
        raise PivotImportError(f"La hoja '{HOJA_EQUIPOS}' está vacía.") from e

    requeridas = {"codigo", "nombre", "hoja_filtros"}
    faltantes = requeridas - set(header)
    if faltantes:
        raise PivotImportError(
            f"A la hoja '{HOJA_EQUIPOS}' le faltan las columnas {sorted(faltantes)}. "
            f"Encontradas: {header}"
        )

    indice = {nombre: header.index(nombre) for nombre in header if nombre}
    equipos: list[dict[str, str | bool]] = []
    for fila in filas:
        codigo = str(fila[indice["codigo"]] or "").strip()
        if not codigo or codigo.lower() in ("nan", "none"):
            continue
        activo_raw = (
            str(fila[indice["activo"]] or "true").strip().lower() if "activo" in indice else "true"
        )
        equipos.append(
            {
                "codigo": codigo.upper(),
                "nombre": str(fila[indice["nombre"]] or codigo).strip(),
                "hoja_filtros": str(fila[indice["hoja_filtros"]] or "").strip(),
                "descripcion": (
                    str(fila[indice["descripcion"]] or "").strip()
                    if "descripcion" in indice
                    else ""
                ),
                "activo": activo_raw not in ("false", "0", "no", ""),
            }
        )
    return equipos


def _leer_reglas_de_hoja(
    wb: openpyxl.Workbook, nombre_hoja: str, codigo_equipo: str
) -> list[tuple[str, str, str]]:
    """Devuelve tuplas (tipo, campo, texto_normalizado) de la hoja de filtros de un equipo."""
    if nombre_hoja not in wb.sheetnames:
        raise PivotImportError(
            f"El equipo '{codigo_equipo}' apunta a la hoja '{nombre_hoja}' pero esa hoja "
            f"no existe en el PIVOT. Corrige la columna hoja_filtros en '{HOJA_EQUIPOS}'."
        )
    ws = wb[nombre_hoja]
    max_columna = max(MAPEO_COLUMNAS.keys()) + 1
    vistas: set[tuple[str, str, str]] = set()
    reglas: list[tuple[str, str, str]] = []
    for fila in ws.iter_rows(
        min_row=FILA_INICIO_DATOS_FILTROS, max_col=max_columna, values_only=True
    ):
        for col_idx, (tipo, campo) in MAPEO_COLUMNAS.items():
            valor = fila[col_idx] if col_idx < len(fila) else None
            texto = normalizar_texto(valor)
            if not texto:
                continue
            clave = (tipo, campo, texto)
            if clave in vistas:
                continue
            vistas.add(clave)
            reglas.append(clave)
    return reglas


def _leer_intencion_global(wb: openpyxl.Workbook) -> list[tuple[str, str]]:
    """Devuelve tuplas (tipo, texto) de la hoja de intención global. Hoja opcional."""
    if HOJA_INTENCION_GLOBAL not in wb.sheetnames:
        return []
    ws = wb[HOJA_INTENCION_GLOBAL]
    filas = ws.iter_rows(values_only=True)
    try:
        header = [str(c or "").strip().lower().replace(" ", "_") for c in next(filas)]
    except StopIteration:
        return []

    mapeo = {
        "intencion_requerida": PalabraIntencion.Tipo.REQUERIDA,
        "intencion_vetada": PalabraIntencion.Tipo.VETADA,
    }
    columnas = {header.index(col): tipo for col, tipo in mapeo.items() if col in header}
    vistas: set[tuple[str, str]] = set()
    palabras: list[tuple[str, str]] = []
    for fila in filas:
        for col_idx, tipo in columnas.items():
            valor = fila[col_idx] if col_idx < len(fila) else None
            texto = normalizar_texto(valor)
            if not texto:
                continue
            clave = (tipo, texto)
            if clave in vistas:
                continue
            vistas.add(clave)
            palabras.append(clave)
    return palabras


def importar_pivot(ruta_pivot: Path) -> ResultadoImportacion:
    """Importa el PIVOT completo a la BD. Idempotente: re-ejecutar no duplica (O4).

    El llamador (management command) es responsable de la transacción y del
    registro en EjecucionPipeline; esta función solo sabe de PIVOT y modelos.
    """
    if not ruta_pivot.exists():
        raise PivotImportError(
            f"No se encontró el PIVOT en '{ruta_pivot}'. Verifica la ruta entregada "
            f"al comando importar_pivot."
        )

    wb = openpyxl.load_workbook(ruta_pivot, read_only=True, data_only=True)
    try:
        resultado = ResultadoImportacion()

        for equipo in _leer_equipos(wb):
            perfil, creado = PerfilFiltro.objects.update_or_create(
                codigo=equipo["codigo"],
                defaults={
                    "nombre": equipo["nombre"],
                    "descripcion": equipo["descripcion"],
                    "activo": equipo["activo"],
                },
            )
            if creado:
                resultado.perfiles_creados += 1
            else:
                resultado.perfiles_actualizados += 1

            reglas = _leer_reglas_de_hoja(wb, str(equipo["hoja_filtros"]), perfil.codigo)
            if not any(tipo == Tipo.INCLUIR for tipo, _, _ in reglas):
                resultado.advertencias.append(
                    f"El equipo '{perfil.codigo}' no tiene reglas de inclusión: "
                    f"su filtrado no producirá resultados hasta que se agreguen."
                )
            for tipo, campo, texto in reglas:
                _, creada = ReglaKeyword.objects.get_or_create(
                    perfil=perfil, tipo=tipo, campo=campo, texto=texto
                )
                if creada:
                    resultado.reglas_creadas += 1
                else:
                    resultado.reglas_existentes += 1

        for tipo, texto in _leer_intencion_global(wb):
            _, creada = PalabraIntencion.objects.get_or_create(tipo=tipo, texto=texto)
            if creada:
                resultado.intencion_creadas += 1

        return resultado
    finally:
        wb.close()
